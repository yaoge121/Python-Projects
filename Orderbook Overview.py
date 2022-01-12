import pandas as pd
import openpyxl
import csv
import numpy as np
from openpyxl.styles import Font,Alignment, Border, Side, Color,NamedStyle
from datetime import datetime,timedelta,date

#TBD-Order
#TBD-order as base
TBD_order_wb1=pd.read_excel("Mapping-test.xlsx", "Orders - EU")
TBD_order_wb1["PGID"]=TBD_order_wb1["Pallet Group ID"]
TBD_order_wb1["TBD ETA on port"]=TBD_order_wb1["ETA on port"]

#TBD filters
TBD_order_filter1= TBD_order_wb1.query("`TBD ETA on port`>= '2021-06-01'")
TBD_order_filter2= TBD_order_filter1[TBD_order_filter1["SF Line Item"].notnull()]
TBD_order_filter2["SF Line Item"] = TBD_order_filter2["SF Line Item"].astype(int)
TBD_order_filter2["SO-item"]= TBD_order_filter2["SO No."].map(str) + "-" + TBD_order_filter2["SF Line Item"].map(str)

TBD_order_wb2=TBD_order_filter2[["PGID", "SAP ID & Item No", "Module type", "Material code",
         "Material description","PCS","Total Watts (MW)","TBD ETA on port","Sales Manager",
         "Customer name","Customer Reference Number","Order No.","SO-item","SLOC"]]
TBD_order=TBD_order_wb2.loc[(TBD_order_wb2["TBD ETA on port"] >= "2021-06-01")]

#TBD Vlookup with 551
zsd0551_order=pd.read_excel("Mapping-test.xlsx", "ZSD0551")
zsd0551_order["IPO"]= zsd0551_order["Sales Order No"].map(str) + "-" + zsd0551_order["Sales Order Item"].map(str)

#merge ATA/ETA log/ETA HQ in one column

order_conditions = [
    (pd.notnull(zsd0551_order["ETA Date(Actual)"])),
    (pd.isnull(zsd0551_order["ETA Date(Actual)"])) & (pd.notnull(zsd0551_order["ETA Date"])),
    (pd.isnull(zsd0551_order["ETA Date(Actual)"])) & (pd.isnull(zsd0551_order["ETA Date"])) & (pd.notnull(zsd0551_order["ETA Date(HQ)"])),
    (pd.isnull(zsd0551_order["ETA Date(Actual)"])) & (pd.isnull(zsd0551_order["ETA Date"])) & (pd.isnull(zsd0551_order["ETA Date(HQ)"])) & (pd.notnull(zsd0551_order["PO Require Date"]))
   ]

order_ETA_values1 = ["ATA", "ETA Log", "ETA HQ", "PO require date"]
order_ETA_values2 = [
    zsd0551_order["ETA Date(Actual)"].map(str),
    zsd0551_order["ETA Date"].map(str),
    zsd0551_order["ETA Date(HQ)"].map(str),
    zsd0551_order["PO Require Date"].map(str)
]

zsd0551_order["SAP ETA category"] = np.select(order_conditions, order_ETA_values1, default=np.nan)
zsd0551_order["SAP ETA on port"] = np.select(order_conditions, order_ETA_values2, default=np.nan)
zsd0551_order["SAP ETA on port"] =pd.to_datetime(zsd0551_order["SAP ETA on port"].map(str), errors='coerce')
zsd0551_order["SAP WH ETA"]= zsd0551_order["SAP ETA on port"] + timedelta(days=10)

#ZSD0551 filters
order_filter1 = zsd0551_order[zsd0551_order["HQ Fulfill Flag"].isnull() & zsd0551_order["Reject Reason Text"].isnull()]
order_filter2 =order_filter1.query("Status=='approved' or Status.isnull()")
zsd0551_order_filter=order_filter2.query("`Sales Type`=='Module' and `PO Require Date`>'2021-01-01'")

zsd0551_order_1=zsd0551_order_filter[["IPO", "Pallet Group ID", "ETA Date(Actual)", "ETA Date", "ETA Date(HQ)", "SAP ETA category",
                          "SAP ETA on port","SAP WH ETA"]]
mergeTBD_order_551=pd.merge(TBD_order, zsd0551_order_1, left_on="SAP ID & Item No", right_on="IPO", how="left")

#merge & vlookup with va05
va05=pd.read_excel("Mapping-test.xlsx","VA05")

#VA05 filters
va05_filter= va05[va05["Item"].notnull()]
va05_filter["Item"] = va05_filter["Item"].astype(int)
va05_filter["SO_item"]=va05_filter["Sales Document"].map(str)+"-"+va05_filter["Item"].map(str)

va05_1=va05_filter[["SO_item","Delivery Date","Delivery Status Description"]]
mergeTBD_551_VA05=pd.merge(mergeTBD_order_551, va05_1, left_on="SO-item", right_on="SO_item", how="left")

#Merge & Vlookup with 0608
zsd0608=pd.read_excel("Mapping-test.xlsx","ZSD0608")
zsd0608["SOitem"]=zsd0608["Sales Document"].map(str)+"-"+zsd0608["Item"].map(str)

#zsd0608 filters
zsd0608_filter =zsd0608.query("Status=='approved' or Status.isnull()")
zsd0608_filter2=zsd0608_filter.query("`Name 1`!='SACS Planning & Procurement'&`Name 1`!='GENL Planning & Procurement'")
zsd0608_filter3=zsd0608_filter2.query("Plant=='GEGB' or Plant=='GEHS' or Plant=='GENL' or Plant=='GEPT' or Plant=='HKCS'or Plant=='SACS'")

zsd0608_1=zsd0608_filter3[["SOitem","Order Material","Order quantity","Allocted quantity",
                  "Unallocated quantity","Undelivery quantity","ReqDlv.dat"]]
mergeall_order=pd.merge(mergeTBD_551_VA05, zsd0608_1, left_on="SO-item", right_on="SOitem", how="left")

overview_order=mergeall_order[["SAP ID & Item No", "Pallet Group ID", "SLOC", "Module type", "Material code", "Material description", "PCS", "Total Watts (MW)",
                 "TBD ETA on port","SAP ETA category","SAP ETA on port","SAP WH ETA","Sales Manager","Customer name","Customer Reference Number","Order No.",
                 "SO-item","Delivery Date","Delivery Status Description","Order quantity","Allocted quantity","Unallocated quantity","Undelivery quantity"]]

overview_order.to_csv("mid-file-order.csv", index=True)

#-----------------------------------------------------------------------------------------------------------------------
#TBD-ATP
# TBD as base
TBD_wb1=pd.read_excel("Mapping-test.xlsx", "TBD-EU")
TBD_wb1["PGID"] = TBD_wb1["Pallet Group ID"].astype(str)
TBD_wb1["TBD ETA on port"] = TBD_wb1["ETA on port"]
TBD_wb1["TBD ETA on port"]= pd.to_datetime(TBD_wb1["TBD ETA on port"], errors='coerce')
TBD_wb1["TBD WH ETA"]= TBD_wb1["TBD ETA on port"] + timedelta(days=10)
TBD_wb1["TBD Module type"]=TBD_wb1["Module type"]
# TBD filters
TBD = TBD_wb1[["ATP", "PGID", "SAP ID & Item No", "Material code", "Material description", "TBD Module type", "PCS", "Total Watts (MW)", "TBD ETA on port", "TBD WH ETA", "Sales Manager",
           "Customer name", "Customer Reference Number", "Order No.", "Comments", "SO No.", "Waiting list 1",
           "Waiting list 2","Reservation Date", "Expiration Date"]]

# TBD Vlookup with 551
zsd0551_ATP = pd.read_excel("Mapping-test.xlsx", "ZSD0551")
zsd0551_ATP["IPO"] = zsd0551_ATP["Sales Order No"].map(str) + "-" + zsd0551_ATP["Sales Order Item"].map(str)

# merge ATA/ETA log/ETA HQ in one column
atp_conditions = [
    (pd.notnull(zsd0551_ATP["ETA Date(Actual)"])),
    (pd.isnull(zsd0551_ATP["ETA Date(Actual)"])) & (pd.notnull(zsd0551_ATP["ETA Date"])),
    (pd.isnull(zsd0551_ATP["ETA Date(Actual)"])) & (pd.isnull(zsd0551_ATP["ETA Date"])) & (pd.notnull(zsd0551_ATP["ETA Date(HQ)"])),
    (pd.isnull(zsd0551_ATP["ETA Date(Actual)"])) & (pd.isnull(zsd0551_ATP["ETA Date"])) & (
        pd.isnull(zsd0551_ATP["ETA Date(HQ)"])) & (pd.notnull(zsd0551_ATP["PO Require Date"]))
]

ETA_values1 = ["ATA", "ETA Log", "ETA HQ", "PO require date"]
ETA_values2 = [
    zsd0551_ATP["ETA Date(Actual)"].map(str),
    zsd0551_ATP["ETA Date"].map(str),
    zsd0551_ATP["ETA Date(HQ)"].map(str),
    zsd0551_ATP["PO Require Date"].map(str)
]

zsd0551_ATP["SAP ETA category"] = np.select(atp_conditions, ETA_values1, default=np.nan)
zsd0551_ATP["SAP ETA on port"] = np.select(atp_conditions, ETA_values2, default=np.nan)
zsd0551_ATP["SAP ETA on port"] = pd.to_datetime(zsd0551_ATP["SAP ETA on port"], errors='coerce')
zsd0551_ATP["SAP WH ETA"] = zsd0551_ATP["SAP ETA on port"] + timedelta(days=10)

# ZSD0551 filters
atp_filter1 = zsd0551_ATP[zsd0551_ATP["HQ Fulfill Flag"].isnull() & zsd0551_ATP["Reject Reason Text"].isnull()]
atp_filter2 = atp_filter1.query("Status=='approved' or Status.isnull()")
zsd0551_atp_filter2 = atp_filter2.query("`Sales Type`=='Module' and `PO Require Date`>'2021-01-01'")

zsd0551_atp_1 = zsd0551_atp_filter2[["IPO", "Material", "Material Desc.", "Module Type", "Power", "Module Color", "Frame Type",
                            "Connector Type","Pallet Group ID", "SAP ETA category", "SAP ETA on port", "SAP WH ETA",
                            "Production Factory", "Incoterms Location"]]
mergeTBD_atp_551 = pd.merge(TBD, zsd0551_atp_1, left_on="SAP ID & Item No", right_on="IPO", how="left")


#WH ETA combine SAP&TBD
ETA_conditions2 = [
    (pd.notnull(mergeTBD_atp_551["SAP WH ETA"])),
    (pd.isnull(mergeTBD_atp_551["SAP WH ETA"])) & (pd.notnull(mergeTBD_atp_551["TBD WH ETA"]))
]

ETA_values3 = [mergeTBD_atp_551["SAP ETA category"], "TBD ETA"]
ETA_values4 = [
    mergeTBD_atp_551["SAP WH ETA"].map(str),
    mergeTBD_atp_551["TBD WH ETA"].map(str),
]
mergeTBD_atp_551["ETA category"] = np.select(ETA_conditions2, ETA_values3, default=np.nan)
mergeTBD_atp_551["WH ETA"] = np.select(ETA_conditions2, ETA_values4, default=np.nan)
mergeTBD_atp_551["WH ETA"]=pd.to_datetime(mergeTBD_atp_551["WH ETA"], errors='coerce')

# Aging formula
mergeTBD_atp_551["Today"]= date.today()
d2=pd.to_datetime(mergeTBD_atp_551["Today"], errors='coerce')
d1=mergeTBD_atp_551["WH ETA"]
mergeTBD_atp_551["Aging days"] = d2 - d1
mergeTBD_atp_551["Aging days"]=mergeTBD_atp_551["Aging days"].apply(lambda x: x.days)

aging_conditions3 = [
    (mergeTBD_atp_551["Module Type"] == "CS6PH-P"),
    (mergeTBD_atp_551["Aging days"] < 0) & (mergeTBD_atp_551["Module Type"] != "CS6PH-P"),
    (mergeTBD_atp_551["Aging days"] >= 0) & (mergeTBD_atp_551["Aging days"] < 30),
    (mergeTBD_atp_551["Aging days"] >= 30) & (mergeTBD_atp_551["Aging days"] < 45),
    (mergeTBD_atp_551["Aging days"] >= 45) & (mergeTBD_atp_551["Aging days"] < 60),
    (mergeTBD_atp_551["Aging days"] >= 60)
]

aging_values5 = ["Classic Modules","Not Aging","Below 30 days","30-45 days","Imminent Aging","Aging above 60 days"]

mergeTBD_atp_551["Aging Category"] = np.select(aging_conditions3, aging_values5, default=np.nan)

#EOM aging
def last_day_of_month(ds):
    cur_ds = datetime.strptime(ds, '%Y-%m-%d')
    extrayear, month = divmod(cur_ds.month, 12)
    next_month = datetime(year=cur_ds.year + extrayear, month=month + 1, day=1)
    last_day_month = next_month - timedelta(days=1)
    return datetime.strftime(last_day_month, '%Y-%m-%d')


EOM_date_str=last_day_of_month(date.today().strftime('%Y-%m-%d'))
EOM_date=datetime.strptime(EOM_date_str, '%Y-%m-%d')
today_date=datetime.strptime(date.today().strftime('%Y-%m-%d'), '%Y-%m-%d')
Datediff=(EOM_date-today_date).days
EOM_std=60-int(Datediff)

EOM_conditions4 = [
    (mergeTBD_atp_551["Module Type"] == "CS6PH-P"),
    (mergeTBD_atp_551["Aging days"] <= EOM_std) & (mergeTBD_atp_551["Module Type"] != "CS6PH-P"),
    (mergeTBD_atp_551["Aging days"] > EOM_std) & (mergeTBD_atp_551["Module Type"] != "CS6PH-P"),
]

EOM_values6 = ["Classic Modules","No EOM aging risk","EOM aging risk"]

mergeTBD_atp_551["EOM Aging Category"] = np.select(EOM_conditions4, EOM_values6, default=np.nan)

#ZSD0150 product tech details
zsd0150= pd.read_excel("Mapping-test.xlsx", "ZSD0150")
#groupby
gb= zsd0150.groupby("Pallet Group ID")
gb1=gb.first()
zsd0150_1=gb1[["Module Appearance Color","Power Enhancer #1(Y/N)","Cell Bus Bar","Module Manufacturer",
               "Cell Patterns","MCS Cert.","UNI9177 Cert."]]
mergeTBD_551_150 = pd.merge(mergeTBD_atp_551, zsd0150_1, left_on="Pallet Group ID", right_on="Pallet Group ID", how="left")

overview_tbd=mergeTBD_551_150[["ATP", "PGID", "Material code", "Material description", "TBD Module type", "PCS", "Total Watts (MW)", "TBD WH ETA", "Sales Manager",
                           "Customer name", "Customer Reference Number", "Order No.", "Comments", "SO No.", "Waiting list 1",
                           "Waiting list 2","Reservation Date", "Expiration Date","IPO", "Material", "Material Desc.", "Module Type", "Power", "Module Color", "Frame Type",
                            "Connector Type","Pallet Group ID", "ETA category", "SAP ETA on port", "WH ETA","Aging days","Aging Category","EOM Aging Category",
                           "Production Factory", "Incoterms Location","Module Appearance Color","Power Enhancer #1(Y/N)","Cell Bus Bar","Module Manufacturer",
                           "Cell Patterns","MCS Cert.","UNI9177 Cert."]]
overview_tbd.to_csv("mid-file-tbd.csv", index=True)


#CSV transfer to excel
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title="New TBD"
ws2 = wb.create_sheet(title="Order")

with open("mid-file-order.csv") as file:
    reader = csv.reader(file, delimiter=',')
    for row in reader:
        ws2.append(row)

with open("mid-file-tbd.csv") as file:
    reader = csv.reader(file, delimiter=',')
    for row in reader:
        ws1.append(row)

# create a style template for the header row
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")
header.color=Color(indexed=0)

# apply this to all first row (header) cells
for ws in wb:
    header_row = ws[1]
    for cell in header_row:
        cell.style = header
    ws["A1"]="S No"

wb.save("EMEA fulfill overview-test.xlsx")